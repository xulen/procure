"""
Módulo de generación de reportes amigables para usuarios finales.

Genera dos formatos a partir del índice (_index.json):
  1. HTML estático (_report.html) — tabla limpia con enlaces a carpetas locales
  2. Excel (_report.xlsx) — una hoja por fuente, columnas organizadas

Uso:
    python src/report.py --output-dir nachus   # Genera reportes en nachus/
    python src/report.py --output-dir nachus --format html  # Solo HTML
    python src/report.py --output-dir nachus --format xlsx  # Solo Excel
"""

import argparse
import json
import os
import sys
import glob as glob_module
from datetime import datetime, timezone


# ─── Descubrimiento de fuentes ────────────────────────────────────────────────


def discover_sources(root_project_dir):
    """
    Descubre todas las fuentes (directorios con _index.json) bajo root_project_dir.

    Args:
        root_project_dir: Directorio raíz del proyecto (ej. /home/xulen/proyectos/procurment)

    Returns:
        lista de dicts {'name': str, 'path': str, 'index': dict}
    """
    sources = []

    # Buscar todos los _index.json en subdirectorios del root
    pattern = os.path.join(root_project_dir, "**", "_index.json")
    index_files = glob_module.glob(pattern, recursive=True)

    for index_path in index_files:
        # Extraer el nombre de la fuente (nombre del directorio padre)
        source_dir = os.path.dirname(index_path)
        source_name = os.path.basename(source_dir)

        try:
            with open(index_path, "r", encoding="utf-8") as f:
                index_data = json.load(f)
        except (json.JSONDecodeError, IOError) as err:
            print(f"  ⚠ No se pudo leer {index_path}: {err}")
            continue

        sources.append({
            "name": source_name,
            "path": source_dir,
            "index": index_data,
        })

    return sources


# ─── Construcción de datos para reportes ────────────────────────────────────────


def build_report_data(sources):
    """
    Construye la estructura plana de datos a partir de las fuentes descubiertas.

    Returns:
        lista de dicts con keys:
          - source: nombre de la fuente
          - title, url, closing_date, status, country
          - doc_count: número de documentos descargados
          - local_path: ruta absoluta a la carpeta del proyecto
          - documents: lista de nombres de archivos en la carpeta
    """
    all_projects = []

    for source in sources:
        index_data = source["index"]
        source_path = source["path"]
        source_name = source["name"]

        for slug, info in index_data.get("slugs", {}).items():
            # Buscar la carpeta del proyecto
            project_dir = os.path.join(source_path, slug)
            docs_dir = os.path.join(project_dir, "documentos")
            documents = []
            doc_count = 0

            if os.path.isdir(docs_dir):
                documents = [f for f in os.listdir(docs_dir)
                             if os.path.isfile(os.path.join(docs_dir, f))]
                doc_count = len(documents)

            all_projects.append({
                "source": source_name,
                "title": info.get("title", slug),
                "url": info.get("url", ""),
                "closing_date": info.get("closing_date"),
                "status": info.get("status"),
                "country": info.get("country"),
                "doc_count": doc_count,
                "local_path": project_dir,
                "documents": sorted(documents),
            })

    return all_projects


# ─── Generador HTML ────────────────────────────────────────────────────────────


def generate_html(report_data, output_path):
    """
    Genera un archivo HTML estático con tabla de proyectos y enlaces a carpetas.

    Args:
        report_data: lista de dicts retornada por build_report_data()
        output_path: ruta donde escribir el HTML
    """
    # Agrupar por fuente
    by_source = {}
    for project in report_data:
        by_source.setdefault(project["source"], []).append(project)

    total_docs = sum(p["doc_count"] for p in report_data)
    total_sources = len(by_source)
    total_projects = len(report_data)

    # Generar filas de la tabla por fuente
    rows_html = ""
    for source_name, projects in sorted(by_source.items()):
        rows_html += f'    <tr class="source-header"><td colspan="7">📂 {source_name} ({len(projects)} proyectos)</td></tr>\n'
        for p in projects:
            status_badge = _status_badge(p["status"])
            country_cell = f"<span class='country'>{_escape(p['country'])}</span>" if p["country"] else '<span class="country">—</span>'
            closing_cell = _escape(p["closing_date"]) if p["closing_date"] else "—"
            docs_cell = f'<span class="docs">{p["doc_count"]} doc(s)</span>' if p["doc_count"] > 0 else '<span class="docs">0</span>'

            # Lista de documentos como tooltip (title attribute)
            doc_titles = "; ".join(p["documents"]) if p["documents"] else ""
            docs_link = f'<a href="{_escape(p["local_path"])}" title="{_escape(doc_titles)}" class="docs-link">{docs_cell}</a>' if p["doc_count"] > 0 else docs_cell

            # Link a la carpeta del proyecto (abre documentos)
            project_link = f'<a href="{_escape(p["local_path"])}" class="project-link">{_escape(p["title"])}</a>'

            rows_html += f"""    <tr>
      <td class="source-tag">{_escape(source_name)}</td>
      <td>{project_link}</td>
      <td>{country_cell}</td>
      <td>{closing_cell}</td>
      <td>{status_badge}</td>
      <td>{docs_link}</td>
      <td><a href="{_escape(p["local_path"])}" class="folder-link">📁 Abrir</a></td>
    </tr>\n"""

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Reporte de Convocatorias — Procurment</title>
  <style>
    :root {{
      --bg: #f8f9fa;
      --surface: #ffffff;
      --border: #dee2e6;
      --text: #212529;
      --text-muted: #6c757d;
      --accent: #0d6efd;
      --accent-light: #e7f1ff;
      --green: #198754;
      --green-light: #d1e7dd;
      --red: #dc3545;
      --red-light: #f8d7da;
      --yellow: #ffc107;
      --yellow-light: #fff3cd;
    }}

    * {{ box-sizing: border-box; margin: 0; padding: 0; }}

    body {{
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
      background: var(--bg);
      color: var(--text);
      line-height: 1.5;
      padding: 2rem;
    }}

    .container {{
      max-width: 1400px;
      margin: 0 auto;
    }}

    header {{
      margin-bottom: 2rem;
    }}

    header h1 {{
      font-size: 1.75rem;
      font-weight: 700;
      margin-bottom: 0.5rem;
    }}

    header p {{
      color: var(--text-muted);
      font-size: 0.95rem;
    }}

    .stats {{
      display: flex;
      gap: 1rem;
      margin-bottom: 2rem;
      flex-wrap: wrap;
    }}

    .stat-card {{
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 1rem 1.5rem;
      min-width: 160px;
      flex: 1;
    }}

    .stat-card .label {{
      font-size: 0.8rem;
      color: var(--text-muted);
      text-transform: uppercase;
      letter-spacing: 0.5px;
    }}

    .stat-card .value {{
      font-size: 1.5rem;
      font-weight: 700;
      margin-top: 0.25rem;
    }}

    table {{
      width: 100%;
      border-collapse: collapse;
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: 8px;
      overflow: hidden;
      font-size: 0.9rem;
    }}

    thead th {{
      background: #f1f3f5;
      padding: 0.75rem 1rem;
      text-align: left;
      font-weight: 600;
      color: var(--text-muted);
      font-size: 0.8rem;
      text-transform: uppercase;
      letter-spacing: 0.5px;
      border-bottom: 2px solid var(--border);
    }}

    tbody td {{
      padding: 0.6rem 1rem;
      border-bottom: 1px solid var(--border);
      vertical-align: middle;
    }}

    tbody tr:last-child td {{
      border-bottom: none;
    }}

    .source-header {{
      background: #f8f9fa !important;
    }}

    .source-header td {{
      font-weight: 600;
      padding: 0.5rem 1rem !important;
      color: var(--text-muted);
      border-bottom: 2px solid var(--border) !important;
    }}

    .source-tag {{
      font-size: 0.75rem;
      background: var(--accent-light);
      color: var(--accent);
      padding: 0.2rem 0.5rem;
      border-radius: 4px;
      font-weight: 600;
      white-space: nowrap;
    }}

    .project-link {{
      color: var(--text);
      text-decoration: none;
      font-weight: 500;
    }}

    .project-link:hover {{
      color: var(--accent);
      text-decoration: underline;
    }}

    .folder-link {{
      color: var(--accent);
      text-decoration: none;
      font-size: 0.85rem;
    }}

    .folder-link:hover {{
      text-decoration: underline;
    }}

    .docs-link {{
      color: var(--text);
      text-decoration: none;
    }}

    .country {{
      color: var(--text-muted);
      font-size: 0.85rem;
    }}

    .status-badge {{
      display: inline-block;
      padding: 0.15rem 0.5rem;
      border-radius: 4px;
      font-size: 0.75rem;
      font-weight: 600;
    }}

    .status-abierta {{
      background: var(--green-light);
      color: var(--green);
    }}

    .status-cerrada {{
      background: var(--red-light);
      color: var(--red);
    }}

    .status-unknown, .status-none {{
      background: var(--yellow-light);
      color: #856404;
    }}

    footer {{
      margin-top: 2rem;
      text-align: center;
      color: var(--text-muted);
      font-size: 0.8rem;
    }}

    @media (max-width: 768px) {{
      body {{ padding: 1rem; }}
      .stats {{ flex-direction: column; }}
      table {{ font-size: 0.8rem; }}
      thead th, tbody td {{ padding: 0.4rem 0.5rem; }}
    }}
  </style>
</head>
<body>
  <div class="container">
    <header>
      <h1>📋 Reporte de Convocatorias</h1>
      <p>Generado el {datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M")} UTC — Total de {total_projects} proyectos en {total_sources} fuente(s) con {total_docs} documento(s) descargados</p>
    </header>

    <div class="stats">
      <div class="stat-card">
        <div class="label">Fuentes</div>
        <div class="value">{total_sources}</div>
      </div>
      <div class="stat-card">
        <div class="label">Proyectos</div>
        <div class="value">{total_projects}</div>
      </div>
      <div class="stat-card">
        <div class="label">Documentos</div>
        <div class="value">{total_docs}</div>
      </div>
    </div>

    <table>
      <thead>
        <tr>
          <th>Fuente</th>
          <th>Proyecto / Documento</th>
          <th>País</th>
          <th>Cierre</th>
          <th>Estado</th>
          <th>Docs</th>
          <th>Acceso</th>
        </tr>
      </thead>
      <tbody>
{rows_html}      </tbody>
    </table>

    <footer>
      <p>Procurment — Multi-lateral Institutions Scraper (CAF, BID, ...)</p>
    </footer>
  </div>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

    return output_path


# ─── Generador Excel ────────────────────────────────────────────────────────────


def generate_excel(report_data, output_path):
    """
    Genera un archivo Excel con una hoja por fuente.

    Args:
        report_data: lista de dicts retornada por build_report_data()
        output_path: ruta donde escribir el Excel
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ❌ openpyxl no está instalado. Instalar con: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto

    # Colores y estilos
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # Agrupar por fuente
    by_source = {}
    for project in report_data:
        by_source.setdefault(project["source"], []).append(project)

    column_headers = ["Fuente", "Título", "País", "Fecha Cierre", "Estado",
                      "URL Origen", "Documentos Descargados", "Ruta Local"]

    for source_name, projects in sorted(by_source.items()):
        ws = wb.create_sheet(title=source_name[:31])  # Excel limita a 31 chars

        # Escribir encabezados
        for col_idx, header in enumerate(column_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Escribir datos
        for row_idx, p in enumerate(projects, 2):
            # Fuente
            ws.cell(row=row_idx, column=1, value=p["source"]).font = data_font
            ws.cell(row=row_idx, column=1).alignment = data_align

            # Título
            ws.cell(row=row_idx, column=2, value=p["title"]).font = data_font
            ws.cell(row=row_idx, column=2).alignment = data_align

            # País
            country_cell = ws.cell(row=row_idx, column=3, value=p["country"] or "—")
            country_cell.font = data_font
            country_cell.alignment = data_align

            # Fecha cierre
            closing_cell = ws.cell(row=row_idx, column=4, value=p["closing_date"] or "—")
            closing_cell.font = data_font
            closing_cell.alignment = data_align

            # Estado (con color)
            status = p["status"] or "unknown"
            status_cell = ws.cell(row=row_idx, column=5, value=status.capitalize() if status != "unknown" else "—")
            status_cell.font = data_font
            status_cell.alignment = Alignment(horizontal="center", vertical="top")

            if status == "abierta":
                status_cell.fill = green_fill
            elif status == "cerrada":
                status_cell.fill = red_fill
            else:
                status_cell.fill = yellow_fill

            # URL origen
            ws.cell(row=row_idx, column=6, value=p["url"] or "—").font = data_font
            ws.cell(row=row_idx, column=6).alignment = data_align

            # Documentos descargados (lista separada por coma)
            docs_text = ", ".join(p["documents"]) if p["documents"] else "—"
            ws.cell(row=row_idx, column=7, value=docs_text).font = data_font
            ws.cell(row=row_idx, column=7).alignment = data_align

            # Ruta local
            ws.cell(row=row_idx, column=8, value=p["local_path"]).font = data_font
            ws.cell(row=row_idx, column=8).alignment = data_align

        # Auto-adjust column widths
        col_widths = [12, 50, 15, 15, 10, 60, 40, 50]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    # Guardar
    wb.save(output_path)
    return output_path


# ─── Utilidades ──────────────────────────────────────────────────────────────────


def _escape(text):
    """Escapa caracteres especiales para HTML."""
    if not text:
        return ""
    return (str(text)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&#39;"))


def _status_badge(status):
    """Genera un badge HTML para el estado."""
    if not status:
        return '<span class="status-badge status-unknown">—</span>'
    cls = f"status-{status}"
    label = status.capitalize()
    return f'<span class="status-badge {cls}">{label}</span>'


# ─── CLI ────────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="Genera reportes amigables (HTML + Excel) desde el índice de convocatorias.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python src/report.py --output-dir nachus                    # HTML + Excel en nachus/
  python src/report.py --output-dir nachus --format html      # Solo HTML
  python src/report.py --output-dir nachus --format xlsx      # Solo Excel
  python src/report.py --project-root /home/xulen/proyectos/procurment
        """,
    )

    parser.add_argument(
        "--output-dir", "-o",
        type=str,
        help="Directorio de salida donde se generan los reportes (default: detecta automáticamente)",
    )
    parser.add_argument(
        "--project-root",
        type=str,
        default=None,
        help="Directorio raíz del proyecto para descubrir fuentes (default: directorio padre de src/)",
    )
    parser.add_argument(
        "--format", "-f",
        choices=["html", "xlsx", "all"],
        default="all",
        help="Formato de reporte a generar (default: all)",
    )

    args = parser.parse_args()

    # Determinar directorio raíz del proyecto
    if args.project_root:
        project_root = args.project_root
    else:
        # Usar el directorio padre de src/
        src_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(src_dir)

    print(f"🔍 Descubriendo fuentes en: {project_root}")

    # Descubrir fuentes
    sources = discover_sources(project_root)
    if not sources:
        print("❌ No se encontró ningún _index.json. Ejecuta el scraper primero.")
        sys.exit(1)

    for s in sources:
        total_projects = len(s["index"].get("slugs", {}))
        print(f"  📂 {s['name']}: {total_projects} proyecto(s)")

    # Construir datos de reportes
    report_data = build_report_data(sources)
    print(f"\n📊 Total proyectos en reporte: {len(report_data)}")

    # Determinar directorio de salida
    if args.output_dir:
        output_dir = args.output_dir
    elif len(sources) == 1:
        output_dir = sources[0]["path"]
    else:
        # Usar el directorio del primer source como base
        output_dir = sources[0]["path"]

    os.makedirs(output_dir, exist_ok=True)

    # Generar reportes según formato solicitado
    if args.format in ("html", "all"):
        html_path = os.path.join(output_dir, "_report.html")
        print(f"\n📄 Generando HTML: {html_path}")
        generate_html(report_data, html_path)
        print(f"  ✅ HTML generado ({_format_size(os.path.getsize(html_path))})")

    if args.format in ("xlsx", "all"):
        xlsx_path = os.path.join(output_dir, "_report.xlsx")
        print(f"\n📊 Generando Excel: {xlsx_path}")
        generate_excel(report_data, xlsx_path)
        print(f"  ✅ Excel generado ({_format_size(os.path.getsize(xlsx_path))})")

    print(f"\n✅ Reportes guardados en: {output_dir}")


def _format_size(size_bytes):
    """Formatea bytes a string legible."""
    for unit in ["B", "KB", "MB", "GB"]:
        if size_bytes < 1024:
            return f"{size_bytes:.0f}{unit}"
        size_bytes /= 1024
    return f"{size_bytes:.1f}TB"


if __name__ == "__main__":
    main()
